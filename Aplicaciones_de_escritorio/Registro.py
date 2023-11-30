from tkinter import *
from tkinter import ttk
import openpyxl
import glob
import pandas as pd

#colores
azul = "#90ADC6"
gris = "#E9EAEC"

#crear pantalla
ventana = Tk()
ventana.title("Ragistro")
ventana.config(bg=azul)

#titulo
panel_titulo = Frame(ventana, bg=azul)
panel_titulo.pack()

etiqueta_titulo = Label(panel_titulo, font=("arial", 25), text="Registro de datos", bg=gris)
etiqueta_titulo.pack()

#todos los registros
panel_campos = Frame(ventana, bg=azul)
panel_campos.pack(side=LEFT)


etiqueta_nombres = Label(panel_campos, font=("arial", 15), text="Nombres: ", bg=azul)
etiqueta_nombres.grid(row=0, column=0, sticky="e")
nombres = StringVar()
campo_nombres = Entry(panel_campos, width=25, textvariable=nombres)
campo_nombres.focus()
campo_nombres.grid(row=0, column=1, pady=5)


etiqueta_apellidos = Label(panel_campos, font=("arial", 15), text="Apellidos: ", bg=azul)
etiqueta_apellidos.grid(row=1, column=0, sticky="e")
apellidos = StringVar()
campo_apellidos = Entry(panel_campos, width=25, textvariable=apellidos)
campo_apellidos.grid(row=1, column=1, pady=5)


etiqueta_edad = Label(panel_campos, font=("arial", 15), text="Edad: ", bg=azul)
etiqueta_edad.grid(row=2, column=0, sticky="e")
edad = IntVar()
campo_edad = Entry(panel_campos, width=25, textvariable=edad)
campo_edad.delete(0)
campo_edad.grid(row=2, column=1, pady=5)


etiqueta_ocupacion = Label(panel_campos, font=("arial", 15), text="Ocupacion: ", bg=azul)
etiqueta_ocupacion.grid(row=3, column=0, sticky="e")
ocupacion = StringVar()
campo_ocupacion = Entry(panel_campos, width=25, textvariable=ocupacion)
campo_ocupacion.grid(row=3, column=1, pady=5)


etiqueta_genero = Label(panel_campos, font=("arial", 15), text="Genero: ", bg=azul)
etiqueta_genero.grid(row=4, column=0, sticky="e")
genero = StringVar()
generos = ["Masculio", "Femenino"]
campo_genero = ttk.Combobox(panel_campos, width=20, textvariable=genero, values=generos)
campo_genero.grid(row=4, column=1, pady=5)


etiqueta_estdo_civ = Label(panel_campos, font=("arial", 15), text="Estado civil: ", bg=azul)
etiqueta_estdo_civ.grid(row=5, column=0, sticky="e")
estado_civ = StringVar()
estados_civ = ["Soltero", "Casado", "Union libre"]
campo_estdo_civ = ttk.Combobox(panel_campos, width=20, textvariable=estado_civ, values=estados_civ)
campo_estdo_civ.grid(row=5, column=1, pady=5)


# crear el excel

# Ruta de la carpeta donde quieres buscar archivos Excel
carpeta = 'C:\\Users\\Deyvid\\OneDrive\\Documents\\Programas\\Aplicaciones_de_escritotio\\Registro_de_datos_personales\\Registro.xlsx'

# Utilizar glob para buscar archivos que coincidan con el patrón
archivos_excel = glob.glob(carpeta)

# Verificar si se encontraron archivos Excel
if not archivos_excel:
    # Crear un nuevo archivo Excel
    archivo_excel = openpyxl.Workbook()

    # Crear una hoja de trabajo en el archivo
    
    hoja = archivo_excel.active 
    hoja.title = "Mi Hoja de Datos"

    #escribir datos 

    hoja['A1'] = "Nombres"
    hoja['B1'] = "Apellidos"
    hoja['C1'] = "Edad"
    hoja['D1'] = "Ocupacion"
    hoja['E1'] = "Genero"
    hoja['F1'] = "Estado civil"

    # Guardar el archivo Excel con un nombre específico
    archivo_excel.save(carpeta)

#panel para mostrar

panel_mostrar = Frame(ventana)
panel_mostrar.pack(side=LEFT)

#boton para eliminar
panel_herramientas = Frame(panel_mostrar)
panel_herramientas.pack()

def eliminar():
    selecion = mostrar.selection()
    
    if selecion:
        indice = mostrar.index(selecion[0])
        print(indice)


        archivo_excel = pd.read_excel(carpeta, sheet_name="Mi Hoja de Datos")
        hoja.delete_rows(indice + 2)
 
        archivo_excel.save(carpeta)
        tabla()

boton_eliminar = Button(panel_herramientas, font=("arial", 10), text="Eliminar", command=eliminar)
boton_eliminar.pack(side=RIGHT)

# busqueda
def buscar():
    # buscar elementos en el excel principal y crear un dataframe con la busqueda
    palabra = campo_busqueda.get()
    archivo = pd.read_excel(carpeta, sheet_name="Mi Hoja de Datos")
    resultado = archivo[archivo["Nombres"].str.contains(palabra, case=False)]

    
    # borrar pantalla
    for item in mostrar.get_children():
        mostrar.delete(item)

    # motrar busqueda en pantalla
    for index, fila in resultado.iterrows():
        tupla = tuple(fila)

        mostrar.insert("", "end", values=tupla)
        print(index)


busqueda = StringVar()
campo_busqueda = Entry(panel_herramientas, textvariable=busqueda)
campo_busqueda.pack(side=LEFT)

boton_busqueda = Button(panel_herramientas, font=("arial", 10), text="Buscar", command=buscar)
boton_busqueda.pack(side=LEFT)

# mostrar la tabla

lista = list(pd.read_excel(carpeta).columns)

mostrar = ttk.Treeview(panel_mostrar, columns=lista, show='headings')
mostrar.pack()

# mostrar elementos de la tabla
def tabla():
    # Configurar las columnas en el Treeview

    archivo_excel = openpyxl.load_workbook(carpeta)
    hoja = archivo_excel["Mi Hoja de Datos"]

    for columna in lista:
        mostrar.heading(columna, text=columna)
        mostrar.column(columna, width=100)

    for item in mostrar.get_children():
        mostrar.delete(item)

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        mostrar.insert("", "end", values=fila)

tabla()


# Conocer donde debe el programa agregar mas elementos 
cont = 1

archivo_excel = openpyxl.load_workbook(carpeta)
hoja = archivo_excel.active

for fila in hoja.iter_rows():
    cont +=1

archivo_excel.save(carpeta)

#ordenar los elementos

def ordenar():
    
    # Leer el archivo de Excel en un DataFrame
    archivo_excel = pd.read_excel(carpeta, sheet_name='Mi Hoja de Datos')

    # Ordenar el DataFrame por la columna "Nombre" en orden ascendente
    archivo_excel_ordenado = archivo_excel.sort_values(by='Nombres', ascending=True)

    # Guardar el DataFrame ordenado en un nuevo archivo Excel
    archivo_excel_ordenado.to_excel(carpeta, sheet_name='Mi Hoja de Datos', index=False)

#boton

def enviar():
    global cont
    archivo_excel = openpyxl.load_workbook(carpeta)
    hoja = archivo_excel.active 

    hoja[f'A{cont}'] = campo_nombres.get()
    hoja[f'B{cont}'] = campo_apellidos.get()
    hoja[f'C{cont}'] = campo_edad.get()
    hoja[f'D{cont}'] = campo_ocupacion.get()
    hoja[f'E{cont}'] = campo_genero.get()
    hoja[f'F{cont}'] = campo_estdo_civ.get()
    
    archivo_excel.save(carpeta)

    
    ordenar()
    tabla()

    cont += 1

    campo_apellidos.delete(0, "end")
    campo_edad.delete(0, "end")
    campo_estdo_civ.delete(0, "end")
    campo_genero.delete(0, "end")
    campo_nombres.delete(0, "end")
    campo_ocupacion.delete(0, "end")
    campo_nombres.focus()

boton_enviar = Button(panel_campos, text="Enviar", command=enviar)
boton_enviar.grid(row=6, column=1, pady=5)

#loop de ventanna
ventana.mainloop()